import os
from openpyxl import Workbook, load_workbook

BASE_DIR = "zakazlar"

def safe(v):
    if v is None:
        return ""
    if hasattr(v, "tzinfo"):
        return v.replace(tzinfo=None)
    return str(v)

def write_to_excel(data):

    agent = safe(data["agent"])
    obyekt = safe(data["obyekt"])[:31]  # Excel sheet max 31 belgi

    os.makedirs(BASE_DIR, exist_ok=True)

    file_path = os.path.join(BASE_DIR, f"{agent}.xlsx")

    # Excel ochish
    if os.path.exists(file_path):
        try:
            wb = load_workbook(file_path)
        except:
            wb = Workbook()
    else:
        wb = Workbook()

    # Sheet olish yoki yaratish
    if obyekt in wb.sheetnames:
        ws = wb[obyekt]
    else:
        ws = wb.create_sheet(obyekt)

        ws.append([
            "Sana",
            "Mijoz",
            "Marka",
            "Mashina",
            "Qogoz",
            "Kuba",
            "Bonus",
            "Umumiy"
        ])

    ws.append([
        safe(data["sana"]),
        safe(data["mijoz"]),
        safe(data["marka"]),
        safe(data["mashina"]),
        safe(data["qogoz"]),
        safe(data["kuba"]),
        safe(data["bonus"]),
        safe(data["umumiy"])
    ])

    wb.save(file_path)


def delete_from_excel(agent, obyekt, start, end):

    file_path = os.path.join(BASE_DIR, f"{agent}.xlsx")
    print("FILE:", file_path)

    if not os.path.exists(file_path):
        print("FAYL TOPILMADI")
        return

    wb = load_workbook(file_path)

    print("SHEETLAR:", wb.sheetnames)
    print("KERAKLI:", obyekt)

    if obyekt not in wb.sheetnames:
        print("SHEET TOPILMADI")
        return

    ws = wb[obyekt]

    print("MAX ROW:", ws.max_row)

    for row in range(ws.max_row, 1, -1):

        sana = ws.cell(row=row, column=1).value
        print("SANA:", sana)

        if sana and start <= sana <= end:
            print("OCHIRILDI:", row)
            ws.delete_rows(row)

    wb.save(file_path)